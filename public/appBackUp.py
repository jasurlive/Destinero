import sys
import os
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QLabel,
    QLineEdit,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QMessageBox,
    QRadioButton,
    QButtonGroup,
)
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt
from openpyxl import load_workbook


def add_to_excel(file_path, sheet_name, name, coords, image_links):
    try:
        # Maintain existing formatting using openpyxl
        workbook = load_workbook(file_path)

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]

        # Add headers if the sheet is empty
        if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
            sheet.append(["Name", "Coords", "Image Links"])

        # Append the new data
        sheet.append([name, coords, image_links])

        # Save the updated workbook
        workbook.save(file_path)
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False


class FuturisticApp(QWidget):
    def __init__(self):
        super().__init__()
        self.file_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "data.xlsx"
        )
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Futuristic Excel Updater")
        self.setStyleSheet("background-color: #2c2f33; color: #ffffff;")

        font = QFont("Arial", 12, QFont.Bold)

        title = QLabel("Add Data to Excel")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)

        name_label = QLabel("Name:")
        name_label.setFont(font)
        self.name_input = QLineEdit()
        self.name_input.setStyleSheet(
            "background-color: #ffffff; color: #000000; border-radius: 5px; padding: 5px;"
        )

        coords_label = QLabel("Coords:")
        coords_label.setFont(font)
        self.coords_input = QLineEdit()
        self.coords_input.setStyleSheet(
            "background-color: #ffffff; color: #000000; border-radius: 5px; padding: 5px;"
        )

        image_links_label = QLabel("Image Links:")
        image_links_label.setFont(font)
        self.image_links_input = QLineEdit()
        self.image_links_input.setStyleSheet(
            "background-color: #ffffff; color: #000000; border-radius: 5px; padding: 5px;"
        )

        sheet_label = QLabel("Select Sheet:")
        sheet_label.setFont(font)
        self.visited_radio = QRadioButton("Visited")
        self.visited_radio.setFont(font)
        self.visited_radio.setStyleSheet("color: #ffffff;")
        self.visited_radio.setChecked(True)

        self.planned_radio = QRadioButton("Planned")
        self.planned_radio.setFont(font)
        self.planned_radio.setStyleSheet("color: #ffffff;")

        self.sheet_group = QButtonGroup()
        self.sheet_group.addButton(self.visited_radio)
        self.sheet_group.addButton(self.planned_radio)

        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(self.visited_radio)
        sheet_layout.addWidget(self.planned_radio)

        add_button = QPushButton("Add to Excel")
        add_button.setFont(font)
        add_button.setStyleSheet(
            "background-color: #7289da; color: #ffffff; border-radius: 10px; padding: 10px;"
        )
        add_button.clicked.connect(self.add_to_excel)

        layout = QVBoxLayout()
        layout.addWidget(title)

        form_layout = QVBoxLayout()
        form_layout.addWidget(name_label)
        form_layout.addWidget(self.name_input)
        form_layout.addWidget(coords_label)
        form_layout.addWidget(self.coords_input)
        form_layout.addWidget(image_links_label)
        form_layout.addWidget(self.image_links_input)
        form_layout.addWidget(sheet_label)
        form_layout.addLayout(sheet_layout)

        layout.addLayout(form_layout)
        layout.addWidget(add_button)

        self.setLayout(layout)

    def add_to_excel(self):
        name = self.name_input.text()
        coords = self.coords_input.text()
        image_links = self.image_links_input.text()

        sheet_name = "Visited" if self.visited_radio.isChecked() else "Planned"

        if not name or not coords or not image_links:
            QMessageBox.warning(self, "Input Error", "All fields must be filled!")
            return

        success = add_to_excel(self.file_path, sheet_name, name, coords, image_links)

        if success:
            QMessageBox.information(self, "Success", "Data added successfully!")
            self.name_input.clear()
            self.coords_input.clear()
            self.image_links_input.clear()
        else:
            QMessageBox.critical(self, "Error", "Failed to add data to Excel.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FuturisticApp()
    window.resize(400, 400)
    window.show()
    sys.exit(app.exec_())
