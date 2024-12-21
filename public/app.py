import sys
import os
from PyQt5.QtWidgets import (
    QApplication,
    QMessageBox,
    QSystemTrayIcon,
    QMenu,
    QAction,
    QMainWindow,
    QStyle,
)
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from design import Design
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler


def add_to_excel(file_path, sheet_name, name, coords, image_links):
    try:
        workbook = load_workbook(file_path)

        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]

        if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
            sheet.append(["Name", "Coords", "Image Links"])

        sheet.append([name, coords, image_links])

        workbook.save(file_path)
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False


class FuturisticApp(Design):
    def __init__(self):
        super().__init__()

        self.file_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "data.xlsx"
        )
        self.add_button.clicked.connect(self.add_to_excel)
        self.open_excel_button.clicked.connect(self.open_excel)

        # Set the window to always stay on top
        self.setWindowFlag(Qt.WindowStaysOnTopHint)

        # Use a built-in icon from PyQt5 (e.g., Desktop Icon)
        self.setWindowIcon(QApplication.style().standardIcon(QStyle.SP_DesktopIcon))

        # Optionally, create a system tray icon (with a standard icon)
        self.create_tray_icon()

        # Move the window to the right side of the screen
        screen_geometry = QApplication.primaryScreen().geometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()

        # Position the window to the right (adjust 50 pixels from the edge)
        self.move(
            screen_width - self.width() - 50, (screen_height - self.height()) // 2
        )

    def add_to_excel(self):
        name = self.name_input.text()
        coords = self.coords_input.text()
        image_links = self.image_links_input.text()

        sheet_name = "Visited" if self.visited_radio.isChecked() else "Planned"

        if not name or not coords or not image_links:
            self.show_message("All fields must be filled!", error=True)
            return

        success = add_to_excel(self.file_path, sheet_name, name, coords, image_links)

        if success:
            self.show_message("Data added successfully!")
            self.name_input.clear()
            self.coords_input.clear()
            self.image_links_input.clear()
        else:
            self.show_message("Failed to add data to Excel.", error=True)

    def show_message(self, message, error=False):
        """
        Display a message on the GUI window.
        :param message: The message to display.
        :param error: If True, show the message as an error (red text).
        """
        self.message_label.setText(message)
        self.message_label.setStyleSheet(
            "color: #ff5555;"
            if error
            else "color: #55ff55;"  # Red for errors, green for success
        )

    def create_tray_icon(self):
        tray_icon = QSystemTrayIcon(
            QApplication.style().standardIcon(QStyle.SP_DesktopIcon), self
        )
        tray_icon.setToolTip("ReactJS Map Editor")

        # Create a menu for the tray icon
        tray_menu = QMenu()

        # Add an action to quit the application
        quit_action = QAction("Quit", self)
        quit_action.triggered.connect(self.quit_app)
        tray_menu.addAction(quit_action)

        # Set the menu and activate the tray icon
        tray_icon.setContextMenu(tray_menu)
        tray_icon.show()

    def quit_app(self):
        QApplication.quit()


class CodeChangeHandler(FileSystemEventHandler):
    def __init__(self, script_path, observer, watch_dir):
        self.script_path = script_path
        self.observer = observer
        self.watch_dir = watch_dir

    def on_modified(self, event):
        # Check if the modified file is a Python file
        if event.src_path.endswith(".py"):
            print(f"Detected changes in the file: {event.src_path}")
            self.restart_app()

    def restart_app(self):
        print("Restarting the application...")
        self.observer.stop()  # Stop the observer before restarting
        python = sys.executable
        os.execl(python, python, *sys.argv)


if __name__ == "__main__":
    script_path = os.path.abspath(__file__)
    watch_dir = os.path.dirname(
        script_path
    )  # The directory where your Python files are stored

    # Setup watchdog to monitor all Python files in the directory for changes
    observer = Observer()
    event_handler = CodeChangeHandler(script_path, observer, watch_dir)
    observer.schedule(
        event_handler, watch_dir, recursive=True
    )  # Watch entire directory (including subdirectories)
    observer.start()

    try:
        app = QApplication(sys.argv)
        window = FuturisticApp()
        window.resize(600, 600)
        window.show()
        sys.exit(app.exec_())
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
